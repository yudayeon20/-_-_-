"""봄내티움 성장 정원 로컬 실행 서버.

실행:
    python app.py

접속:
    http://127.0.0.1:8080
"""

from __future__ import annotations

import base64
import csv
import hashlib
import hmac
import json
import mimetypes
import os
import re
import secrets
import shutil
import threading
import webbrowser
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import urlsplit

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parent
WEB_DIR = PROJECT_DIR / "dist"
PUBLIC_DIR = PROJECT_DIR / "public"
STUDENT_FILE = PROJECT_DIR / "data" / "chuncheon_students.xlsx"
STUDENT_BACKUP = PROJECT_DIR / "data" / "chuncheon_students.backup.xlsx"
HOST = "127.0.0.1"
PORT = int(os.environ.get("BOMNAE_PORT", "8080"))
STUDENT_LOCK = threading.Lock()


def student_rows():
    workbook = load_workbook(STUDENT_FILE, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    headers = [cell.value for cell in sheet[1]]
    students = [dict(zip(headers, row)) for row in sheet.iter_rows(min_row=2, values_only=True)]
    workbook.close()
    return students


def program_rows():
    program_file = PROJECT_DIR / "data" / "programs.csv"
    fields = (
        "program_id", "title", "provider", "summary", "target_grade_codes",
        "grade_label", "is_sw", "is_ai", "is_bio", "is_sw_ai_related",
        "is_bio_related", "interest_tags", "growth_label", "is_online",
        "place_name", "start_date", "end_date", "status_label", "fee_label",
        "detail_url",
    )
    with program_file.open("r", encoding="utf-8-sig", newline="") as file:
        programs = []
        for row in csv.DictReader(file):
            if not any(
                row.get(field) == "1"
                for field in ("is_sw", "is_ai", "is_bio", "is_sw_ai_related", "is_bio_related")
            ):
                continue
            programs.append({field: row.get(field, "") for field in fields})
        return programs


def public_student(student):
    return {key: value for key, value in student.items() if key != "login_password"}


def hash_password(password):
    salt = secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 200_000)
    return f"pbkdf2_sha256$200000${base64.b64encode(salt).decode()}${base64.b64encode(digest).decode()}"


def password_matches(password, stored_password):
    stored = str(stored_password or "")
    if not stored.startswith("pbkdf2_sha256$"):
        return hmac.compare_digest(password, stored)
    try:
        _, iterations, salt, expected = stored.split("$", 3)
        digest = hashlib.pbkdf2_hmac(
            "sha256", password.encode("utf-8"), base64.b64decode(salt), int(iterations)
        )
        return hmac.compare_digest(base64.b64encode(digest).decode(), expected)
    except (TypeError, ValueError):
        return False


def find_student(login_id, login_password):
    normalized_login_id = str(login_id or "").strip().lower()
    password = str(login_password or "")
    for student in student_rows():
        if (
            str(student.get("login_id") or "").strip().lower() == normalized_login_id
            and password_matches(password, student.get("login_password"))
        ):
            return student
    return None


def create_student(payload):
    required = (
        "login_id", "login_password", "stu_name", "gender", "school_level",
        "grade", "district_name", "interest_category",
    )
    if any(str(payload.get(key, "")).strip() == "" for key in required):
        raise ValueError("모든 필수 항목을 입력해주세요.")

    login_id = str(payload["login_id"]).strip()
    password = str(payload["login_password"])
    if not re.fullmatch(r"[A-Za-z0-9_]{4,20}", login_id):
        raise ValueError("아이디는 영문, 숫자, 밑줄을 사용해 4~20자로 입력해주세요.")
    if len(password) < 6:
        raise ValueError("비밀번호는 6자 이상으로 입력해주세요.")

    with STUDENT_LOCK:
        workbook = load_workbook(STUDENT_FILE)
        sheet = workbook.worksheets[0]
        headers = [cell.value for cell in sheet[1]]
        existing_students = [
            dict(zip(headers, row))
            for row in sheet.iter_rows(min_row=2, values_only=True)
        ]
        if any(
            str(student.get("login_id") or "").strip().lower() == login_id.lower()
            for student in existing_students
        ):
            workbook.close()
            raise ValueError("이미 사용 중인 아이디입니다.")

        existing_ids = [str(student.get("stu_id") or "") for student in existing_students]
        numbers = [
            int(match.group(1))
            for value in existing_ids
            if (match := re.fullmatch(r"(?i)stu_(\d+)", value))
        ]
        next_id = f"STU_{max(numbers, default=0) + 1:03d}"
        school_level = str(payload["school_level"]).strip()
        grade_prefix = {"초등학교": "e", "중학교": "m", "고등학교": "h"}
        if school_level not in grade_prefix:
            workbook.close()
            raise ValueError("학교급을 올바르게 선택해주세요.")
        grade = int(payload["grade"])
        max_grade = 6 if school_level == "초등학교" else 3
        if not 1 <= grade <= max_grade:
            workbook.close()
            raise ValueError("학년을 올바르게 선택해주세요.")

        student = {
            "stu_id": next_id,
            "login_id": login_id,
            "login_password": hash_password(password),
            "stu_name": str(payload["stu_name"]).strip(),
            "gender": str(payload["gender"]).strip(),
            "school_level": school_level,
            "grade": f"{grade_prefix[school_level]}{grade}",
            "district_name": str(payload["district_name"]).strip(),
            "interest_category": str(payload["interest_category"]).strip(),
            "growth_stage": "씨앗",
            "fertilizer_count": 0,
            "harvested_fruit_count": 0,
        }
        if not STUDENT_BACKUP.exists():
            shutil.copy2(STUDENT_FILE, STUDENT_BACKUP)
        sheet.append([student.get(header, "") for header in headers])
        workbook.save(STUDENT_FILE)
        workbook.close()
        return public_student(student)


def update_student(payload):
    stu_id = str(payload.get("stu_id") or "").strip().upper()
    allowed_fields = {
        "stu_name", "gender", "school_level", "grade",
        "district_name", "interest_category",
    }
    if not stu_id:
        raise ValueError("학생 ID가 필요합니다.")

    with STUDENT_LOCK:
        workbook = load_workbook(STUDENT_FILE)
        sheet = workbook.worksheets[0]
        headers = [cell.value for cell in sheet[1]]
        stu_id_column = headers.index("stu_id") + 1
        target_row = next(
            (
                row
                for row in range(2, sheet.max_row + 1)
                if str(sheet.cell(row, stu_id_column).value or "").strip().upper() == stu_id
            ),
            None,
        )
        if target_row is None:
            workbook.close()
            raise ValueError("학생 정보를 찾을 수 없습니다.")

        school_level = str(payload.get("school_level") or "").strip()
        grade_value = str(payload.get("grade") or "").strip().lower()
        grade_prefix = {"초등학교": "e", "중학교": "m", "고등학교": "h"}
        if school_level:
            if school_level not in grade_prefix:
                workbook.close()
                raise ValueError("학교급을 올바르게 선택해주세요.")
            number = re.sub(r"^[emh]", "", grade_value)
            grade = int(number)
            max_grade = 6 if school_level == "초등학교" else 3
            if not 1 <= grade <= max_grade:
                workbook.close()
                raise ValueError("학년을 올바르게 선택해주세요.")
            payload["grade"] = f"{grade_prefix[school_level]}{grade}"

        for field in allowed_fields:
            if field not in payload:
                continue
            value = str(payload[field]).strip()
            if not value:
                workbook.close()
                raise ValueError("기본 정보 항목을 모두 입력해주세요.")
            sheet.cell(target_row, headers.index(field) + 1).value = value

        workbook.save(STUDENT_FILE)
        workbook.close()
        updated = next(
            student for student in student_rows()
            if str(student.get("stu_id") or "").strip().upper() == stu_id
        )
        return public_student(updated)


class WebHandler(SimpleHTTPRequestHandler):
    """정적 파일을 제공하고 없는 화면 경로는 메인 페이지로 연결합니다."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(WEB_DIR), **kwargs)

    def do_GET(self):
        request_path = urlsplit(self.path).path
        if request_path == "/api/students":
            return self.send_json({"students": [public_student(student) for student in student_rows()]})
        if request_path == "/api/programs":
            return self.send_json({"programs": program_rows()})

        public_file = (PUBLIC_DIR / request_path.lstrip("/")).resolve()
        if public_file.is_relative_to(PUBLIC_DIR.resolve()) and public_file.is_file():
            return self.send_public_file(public_file)

        local_path = WEB_DIR / request_path.lstrip("/")

        if request_path == "/":
            self.path = "/index.html"
        elif not local_path.exists() and "." not in Path(request_path).name:
            self.path = "/index.html"

        super().do_GET()

    def send_public_file(self, path):
        body = path.read_bytes()
        content_type, _ = mimetypes.guess_type(path.name)
        self.send_response(200)
        self.send_header("Content-Type", content_type or "application/octet-stream")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store, max-age=0")
        self.end_headers()
        self.wfile.write(body)

    def do_POST(self):
        request_path = urlsplit(self.path).path
        if request_path not in ("/api/students", "/api/login", "/api/students/update"):
            return self.send_error(404)
        try:
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            if request_path == "/api/login":
                student = find_student(payload.get("login_id"), payload.get("login_password"))
                if not student:
                    return self.send_json({"error": "아이디 또는 비밀번호가 올바르지 않습니다."}, status=401)
                return self.send_json({"student": public_student(student)})
            if request_path == "/api/students/update":
                student = update_student(payload)
                return self.send_json({"student": student})
            student = create_student(payload)
            self.send_json({"student": student}, status=201)
        except (ValueError, TypeError, json.JSONDecodeError) as error:
            self.send_json({"error": str(error)}, status=400)
        except Exception as error:
            print(f"[봄내티움] 학생 저장 오류: {error}")
            self.send_json({"error": "학생 정보를 저장하지 못했습니다."}, status=500)

    def send_json(self, value, status=200):
        body = json.dumps(value, ensure_ascii=False, default=str).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, format, *args):
        print(f"[봄내티움] {self.address_string()} - {format % args}")


def main():
    index_file = WEB_DIR / "index.html"
    if not index_file.exists():
        raise SystemExit(
            "실행할 웹 파일을 찾지 못했습니다.\n"
            "C:\\chuncheon 폴더에서 먼저 npm run build를 실행해주세요."
        )

    url = f"http://{HOST}:{PORT}"
    server = ThreadingHTTPServer((HOST, PORT), WebHandler)

    print("=" * 52)
    print("  봄내티움 성장 정원이 실행되었습니다.")
    print(f"  접속 주소: {url}")
    print("  종료하려면 이 창에서 Ctrl+C를 누르세요.")
    print("=" * 52)

    if os.environ.get("BOMNAE_NO_BROWSER") != "1":
        threading.Timer(0.8, lambda: webbrowser.open(url)).start()

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n서버를 종료합니다.")
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
